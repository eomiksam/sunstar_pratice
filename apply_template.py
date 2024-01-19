# apply_template.py

from openpyxl import load_workbook
from docx import Document
from data_extraction import extract_data

def apply_template():
    # 데이터 추출
    data = extract_data()

    # 엑셀 템플릿 파일 열기
    wb = load_workbook('template/회생채권자의 목록.xlsx')  # your_template.xlsx 부분을 실제 템플릿 파일 경로로 변경해야 합니다.
    ws = wb.active

    # 엑셀 템플릿 파일에 데이터 적용
    for row, record in enumerate(data, start=2):  # start=2는 두 번째 행부터 시작한다는 의미입니다.
        for col, value in enumerate(record, start=1):  # start=1는 첫 번째 열부터 시작한다는 의미입니다.
            ws.cell(row=row, column=col).value = value

    # 엑셀 템플릿 파일 저장
    wb.save('template/회생채권자의 목록_저장.xlsx')  # your_output.xlsx 부분을 원하는 출력 파일 이름으로 변경해야 합니다.

    # 워드 템플릿 파일 열기
    doc = Document('template/별지 83 회생채권자의 목록.docx')  # your_template.docx 부분을 실제 템플릿 파일 경로로 변경해야 합니다.

    # 워드 템플릿 파일에 데이터 적용
    for paragraph in doc.paragraphs:
        if 'placeholder' in paragraph.text:  # placeholder 부분을 실제 플레이스홀더로 변경해야 합니다.
            for record in data:
                # 플레이스홀더를 데이터로 교체
                paragraph.text = paragraph.text.replace('placeholder', str(record))

    # 워드 템플릿 파일 저장
    doc.save('template/별지 83 회생채권자의 목록_저장.docx')  # your_output.docx 부분을 원하는 출력 파일 이름으로 변경해야 합니다.
